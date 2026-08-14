"""Utilitaires pour la table trppu_site : parsing Excel, normalisation, requêtes SQL."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from pydantic import ValidationError

from app.db.mysql import db_read

from .schemas import BulkUploadError, SiteCreate

EXPECTED_HEADERS = ["co_regate", "lb_regate", "type_site", "co_roc"]
REQUIRED_HEADERS = ["co_regate", "type_site", "co_roc"]


SELECT_SITE_EXISTS_SQL = "SELECT co_regate FROM trppu_site WHERE co_regate = %s"


UPSERT_SQL = (
    "INSERT INTO trppu_site (co_regate, lb_regate, type_site, co_roc) "
    "VALUES (%s, %s, %s, %s) "
    "ON DUPLICATE KEY UPDATE "
    "lb_regate = VALUES(lb_regate), "
    "type_site = VALUES(type_site), "
    "co_roc = VALUES(co_roc)"
)


async def fetch_site_or_404(co_regate: str) -> dict[str, Any]:
    """Retourne le site ou lève un 404. Miroir de `fetch_scenario_or_404`.

    Ne sélectionne que la clé : les appelants qui ont besoin du libellé ou du type
    passent par `SELECT_SITE_SQL` (module routes). Le SQL est redéfini ici plutôt
    qu'importé de `routes.py` pour ne pas créer d'import circulaire.
    """
    row = await db_read.fetch_one(SELECT_SITE_EXISTS_SQL, (co_regate,))
    if not row:
        raise HTTPException(status_code=404, detail=f"Site {co_regate} introuvable.")
    return row


def _normalize_co(value: Any) -> str:
    """co_regate / co_roc : padding gauche avec '0' jusqu'à 6 caractères (Excel rogne les zéros)."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)).zfill(6)
    return str(value).strip().zfill(6)


def parse_excel_sites(content: bytes) -> tuple[list[SiteCreate], list[BulkUploadError]]:
    """Lit un fichier Excel et retourne (sites valides, erreurs par ligne).

    Format attendu : première feuille, ligne 1 = en-têtes, ligne 2+ = données.
    Les en-têtes obligatoires sont : co_regate, type_site, co_roc.
    La colonne lb_regate est optionnelle.
    """
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as e:
        raise ValueError("Fichier Excel vide.") from e

    headers = [
        (str(h).strip().lower() if h is not None else "") for h in header_row
    ]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError(
            f"Colonnes obligatoires manquantes dans l'Excel : {missing}. "
            f"Colonnes attendues : {EXPECTED_HEADERS}"
        )

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS if h in headers}

    valid: list[SiteCreate] = []
    errors: list[BulkUploadError] = []

    for excel_row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None or cell == "" for cell in row):
            continue

        raw = {h: row[i] if i < len(row) else None for h, i in idx.items()}

        try:
            payload = {
                "co_regate": _normalize_co(raw.get("co_regate")),
                "lb_regate": (
                    str(raw["lb_regate"]).strip()
                    if raw.get("lb_regate") not in (None, "")
                    else None
                ),
                "type_site": (
                    str(raw["type_site"]).strip() if raw.get("type_site") is not None else None
                ),
                "co_roc": _normalize_co(raw.get("co_roc")),
            }
            site = SiteCreate.model_validate(payload)
            valid.append(site)
        except ValidationError as e:
            errors.append(
                BulkUploadError(row=excel_row_num, error=str(e), raw=raw)
            )
        except (ValueError, TypeError) as e:
            errors.append(
                BulkUploadError(row=excel_row_num, error=str(e), raw=raw)
            )

    return valid, errors


def site_to_upsert_params(site: SiteCreate) -> tuple:
    """Convertit un SiteCreate en tuple de paramètres pour UPSERT_SQL."""
    return (
        site.co_regate,
        site.lb_regate,
        site.type_site,
        site.co_roc,
    )

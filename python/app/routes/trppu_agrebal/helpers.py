"""Utilitaires pour trppu_agrebal : parsing Excel, normalisation, requêtes SQL."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError

from .schemas import AgrebalCreate, BulkUploadError

EXPECTED_HEADERS = ["id_agrebal", "co_regate", "lb_agrebal", "est_actif"]
REQUIRED_HEADERS = ["id_agrebal", "co_regate"]


UPSERT_SQL = (
    "INSERT INTO trppu_agrebal (id_agrebal, co_regate, lb_agrebal, est_actif) "
    "VALUES (%s, %s, %s, %s) "
    "ON DUPLICATE KEY UPDATE "
    "co_regate = VALUES(co_regate), "
    "lb_agrebal = VALUES(lb_agrebal), "
    "est_actif = VALUES(est_actif)"
)


def _normalize_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) == 1
    if isinstance(value, str):
        v = value.strip().lower()
        if v in {"1", "true", "vrai", "oui", "yes", "y", "o", "actif"}:
            return True
        if v in {"0", "false", "faux", "non", "no", "n", "inactif"}:
            return False
    raise ValueError(f"Booléen invalide : {value!r}")


def _normalize_co_regate(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)).zfill(6)
    return str(value).strip().zfill(6)


def _normalize_id(value: Any) -> int:
    if value is None or value == "":
        raise ValueError("id_agrebal obligatoire")
    if isinstance(value, bool):
        raise ValueError("id_agrebal invalide")
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        return int(value.strip())
    raise ValueError(f"id_agrebal invalide : {value!r}")


def parse_excel_agrebals(content: bytes) -> tuple[list[AgrebalCreate], list[BulkUploadError]]:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as e:
        raise ValueError("Fichier Excel vide.") from e

    headers = [
        (str(h).strip().lower() if h is not None else "") for h in header_row
    ]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError(
            f"Colonnes obligatoires manquantes : {missing}. Attendues : {EXPECTED_HEADERS}"
        )

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS if h in headers}

    valid: list[AgrebalCreate] = []
    errors: list[BulkUploadError] = []

    for excel_row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None or cell == "" for cell in row):
            continue

        raw = {h: row[i] if i < len(row) else None for h, i in idx.items()}

        try:
            payload = {
                "id_agrebal": _normalize_id(raw.get("id_agrebal")),
                "co_regate": _normalize_co_regate(raw.get("co_regate")),
                "lb_agrebal": (str(raw["lb_agrebal"]).strip() if raw.get("lb_agrebal") not in (None, "") else None),
                "est_actif": _normalize_bool(raw.get("est_actif"), default=True),
            }
            agr = AgrebalCreate.model_validate(payload)
            valid.append(agr)
        except ValidationError as e:
            errors.append(BulkUploadError(row=excel_row_num, error=str(e), raw=raw))
        except (ValueError, TypeError) as e:
            errors.append(BulkUploadError(row=excel_row_num, error=str(e), raw=raw))

    return valid, errors


def agrebal_to_upsert_params(a: AgrebalCreate) -> tuple:
    return (a.id_agrebal, a.co_regate, a.lb_agrebal, 1 if a.est_actif else 0)

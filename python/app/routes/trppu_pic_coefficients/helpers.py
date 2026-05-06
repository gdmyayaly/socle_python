"""Utilitaires pour trppu_pic_coefficients : parsing Excel, normalisation, requêtes SQL."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError

from .schemas import BulkUploadError, JourSemaineEnum, PicCoefCreate

EXPECTED_HEADERS = [
    "id_pic_version",
    "co_produit",
    "jour_semaine",
    "dt_effet",
    "dt_fin_effet",
    "coef_dense",
    "coef_faible1",
    "coef_faible2",
]
REQUIRED_HEADERS = [
    "id_pic_version",
    "co_produit",
    "jour_semaine",
    "dt_effet",
    "coef_dense",
    "coef_faible1",
    "coef_faible2",
]


# Upsert sur la natural key (UNIQUE uq_picc) — id_pic_coef est auto-généré.
UPSERT_SQL = (
    "INSERT INTO trppu_pic_coefficients "
    "(id_pic_version, co_produit, jour_semaine, dt_effet, dt_fin_effet, "
    " coef_dense, coef_faible1, coef_faible2) "
    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s) "
    "ON DUPLICATE KEY UPDATE "
    "dt_fin_effet = VALUES(dt_fin_effet), "
    "coef_dense = VALUES(coef_dense), "
    "coef_faible1 = VALUES(coef_faible1), "
    "coef_faible2 = VALUES(coef_faible2)"
)


def _normalize_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Date invalide : {value!r}")


def _normalize_co_produit(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)).zfill(2)
    return str(value).strip().upper().zfill(2)


def _normalize_jour_semaine(value: Any) -> str | None:
    if value is None or value == "":
        return None
    return str(value).strip().upper()


def _normalize_int(value: Any, field: str) -> int:
    if value is None or value == "":
        raise ValueError(f"{field} obligatoire")
    if isinstance(value, bool):
        raise ValueError(f"{field} invalide")
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        return int(value.strip())
    raise ValueError(f"{field} invalide : {value!r}")


def _normalize_float(value: Any, field: str) -> float:
    if value is None or value == "":
        raise ValueError(f"{field} obligatoire")
    if isinstance(value, bool):
        raise ValueError(f"{field} invalide")
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        return float(value.strip().replace(",", "."))
    raise ValueError(f"{field} invalide : {value!r}")


def parse_excel_pic_coefs(
    content: bytes,
) -> tuple[list[PicCoefCreate], list[BulkUploadError]]:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as e:
        raise ValueError("Fichier Excel vide.") from e

    headers = [
        (str(h).strip().lower() if h is not None else "") for h in header_row
    ]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError(
            f"Colonnes obligatoires manquantes : {missing}. Attendues : {EXPECTED_HEADERS}"
        )

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS if h in headers}

    valid: list[PicCoefCreate] = []
    errors: list[BulkUploadError] = []

    for excel_row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None or cell == "" for cell in row):
            continue

        raw = {h: row[i] if i < len(row) else None for h, i in idx.items()}

        try:
            payload = {
                "id_pic_version": _normalize_int(raw.get("id_pic_version"), "id_pic_version"),
                "co_produit": _normalize_co_produit(raw.get("co_produit")),
                "jour_semaine": _normalize_jour_semaine(raw.get("jour_semaine")),
                "dt_effet": _normalize_date(raw.get("dt_effet")),
                "dt_fin_effet": _normalize_date(raw.get("dt_fin_effet")),
                "coef_dense": _normalize_float(raw.get("coef_dense"), "coef_dense"),
                "coef_faible1": _normalize_float(raw.get("coef_faible1"), "coef_faible1"),
                "coef_faible2": _normalize_float(raw.get("coef_faible2"), "coef_faible2"),
            }
            picc = PicCoefCreate.model_validate(payload)
            valid.append(picc)
        except ValidationError as e:
            errors.append(BulkUploadError(row=excel_row_num, error=str(e), raw=raw))
        except (ValueError, TypeError) as e:
            errors.append(BulkUploadError(row=excel_row_num, error=str(e), raw=raw))

    return valid, errors


def pic_coef_to_upsert_params(p: PicCoefCreate) -> tuple:
    return (
        p.id_pic_version,
        p.co_produit,
        p.jour_semaine.value,
        p.dt_effet,
        p.dt_fin_effet,
        p.coef_dense,
        p.coef_faible1,
        p.coef_faible2,
    )


__all__ = [
    "EXPECTED_HEADERS",
    "REQUIRED_HEADERS",
    "UPSERT_SQL",
    "JourSemaineEnum",
    "parse_excel_pic_coefs",
    "pic_coef_to_upsert_params",
]

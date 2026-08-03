"""Utilitaires pour trppu_produit : parsing Excel, normalisation, requêtes SQL."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError

from .schemas import BulkUploadError, ProduitCreate

EXPECTED_HEADERS = [
    "co_produit",
    "lb_produit",
    "dt_desactivation",
    "motif_desactivation",
]
REQUIRED_HEADERS = ["co_produit", "lb_produit"]


UPSERT_SQL = (
    "INSERT INTO trppu_produit "
    "(co_produit, lb_produit, dt_desactivation, motif_desactivation) "
    "VALUES (%s, %s, %s, %s) "
    "ON DUPLICATE KEY UPDATE "
    "lb_produit = VALUES(lb_produit), "
    "dt_desactivation = VALUES(dt_desactivation), "
    "motif_desactivation = VALUES(motif_desactivation)"
)


# Création à la volée d'un produit référencé par un TMH. `ON DUPLICATE KEY UPDATE` no-op
# plutôt qu'un SELECT puis INSERT : idempotent, insensible aux accès concurrents, et ne
# touche jamais le libellé d'un produit déjà saisi par le métier.
ENSURE_PRODUIT_SQL = (
    "INSERT INTO trppu_produit (co_produit, lb_produit) VALUES (%s, %s) "
    "ON DUPLICATE KEY UPDATE co_produit = co_produit"
)


def normalize_co_produit_ref(value: Any) -> str:
    """Normalise un code produit venant d'un référentiel externe (Databricks, TMH).

    Contrairement à `_normalize_co_produit` (import Excel), **pas de `zfill(2)`** : un code
    d'un seul caractère est accepté par le pattern TMH et doit être stocké tel quel, sinon la
    ligne `trppu_tmh` pointerait un `co_produit` différent de celui créé ici.
    """
    return "" if value is None else str(value).strip().upper()


async def ensure_produits_exist(tx, codes, libelles: dict[str, str] | None = None) -> list[str]:
    """Crée dans `trppu_produit` les codes absents, avant tout INSERT sur `trppu_tmh`.

    L'agrégation des trafics est dynamique et pilotée par Databricks : un objet peut apparaître
    dans `g_trppu_trafics_*_3` sans exister côté applicatif. Sans cette création, la FK
    `fk_tmh_produit` casse la transaction et l'API remonte une 500 opaque.

    `libelles` = mapping `{co_produit: lb_produit}` issu de `g_trppu_obj_mapping`, à résoudre
    **avant** l'ouverture de la transaction. `lb_produit` étant NOT NULL, on retombe sur le
    code lui-même quand le mapping ne connaît pas l'objet (cas `PR` / `PPI`) ; le libellé se
    corrige ensuite via `PUT /trppu-api/produits/{co_produit}`.

    Retourne les codes réellement créés.
    """
    libelles = libelles or {}
    crees: list[str] = []
    for code in dict.fromkeys(normalize_co_produit_ref(c) for c in codes):
        if not code:
            continue
        libelle = libelles.get(code) or code
        rowcount = await tx.execute(ENSURE_PRODUIT_SQL, (code, libelle))
        if rowcount == 1:  # 1 = insertion, 0 = déjà présent (UPDATE no-op)
            crees.append(code)
    return crees


def _normalize_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Date invalide : {value!r}")


def _normalize_co_produit(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)).zfill(2)
    return str(value).strip().upper().zfill(2)


def parse_excel_produits(content: bytes) -> tuple[list[ProduitCreate], list[BulkUploadError]]:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as e:
        raise ValueError("Fichier Excel vide.") from e

    headers = [
        (str(h).strip().lower() if h is not None else "") for h in header_row
    ]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError(
            f"Colonnes obligatoires manquantes : {missing}. Colonnes attendues : {EXPECTED_HEADERS}"
        )

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS if h in headers}

    valid: list[ProduitCreate] = []
    errors: list[BulkUploadError] = []

    for excel_row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None or cell == "" for cell in row):
            continue

        raw = {h: row[i] if i < len(row) else None for h, i in idx.items()}

        try:
            payload = {
                "co_produit": _normalize_co_produit(raw.get("co_produit")),
                "lb_produit": (str(raw["lb_produit"]).strip() if raw.get("lb_produit") is not None else ""),
                "dt_desactivation": _normalize_date(raw.get("dt_desactivation")),
                "motif_desactivation": (
                    str(raw["motif_desactivation"]).strip()
                    if raw.get("motif_desactivation") not in (None, "")
                    else None
                ),
            }
            produit = ProduitCreate.model_validate(payload)
            valid.append(produit)
        except ValidationError as e:
            errors.append(BulkUploadError(row=excel_row_num, error=str(e), raw=raw))
        except (ValueError, TypeError) as e:
            errors.append(BulkUploadError(row=excel_row_num, error=str(e), raw=raw))

    return valid, errors


def produit_to_upsert_params(p: ProduitCreate) -> tuple:
    return (
        p.co_produit,
        p.lb_produit,
        p.dt_desactivation,
        p.motif_desactivation,
    )

"""Utilitaires pour trppu_pic_version : parsing Excel, normalisation, requêtes SQL."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError

from .schemas import BulkUploadError, NiveauEnum, PicVersionCreate

EXPECTED_HEADERS = [
    "lb_pic_version",
    "niveau",
    "co_regate",
    "dt_activation",
    "dt_desactivation",
    "motif_desactivation",
    "commentaire",
    "est_par_defaut",
]
REQUIRED_HEADERS = ["niveau", "co_regate", "dt_activation"]


INSERT_SQL = (
    "INSERT INTO trppu_pic_version "
    "(lb_pic_version, niveau, co_regate, dt_activation, dt_desactivation, "
    " motif_desactivation, commentaire, est_par_defaut) "
    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
)


def _normalize_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) == 1
    if isinstance(value, str):
        v = value.strip().lower()
        if v in {"1", "true", "vrai", "oui", "yes", "y", "o"}:
            return True
        if v in {"0", "false", "faux", "non", "no", "n"}:
            return False
    raise ValueError(f"Booléen invalide : {value!r}")


def _normalize_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        s = value.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y",
        ):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    raise ValueError(f"Datetime invalide : {value!r}")


def _normalize_co_regate(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)).zfill(6)
    return str(value).strip().zfill(6)


def _normalize_niveau(value: Any) -> str | None:
    if value is None or value == "":
        return None
    return str(value).strip().upper()


def parse_excel_pic_versions(
    content: bytes,
) -> tuple[list[PicVersionCreate], list[BulkUploadError]]:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as e:
        raise ValueError("Fichier Excel vide.") from e

    headers = [
        (str(h).strip().lower() if h is not None else "") for h in header_row
    ]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError(
            f"Colonnes obligatoires manquantes : {missing}. Attendues : {EXPECTED_HEADERS}"
        )

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS if h in headers}

    valid: list[PicVersionCreate] = []
    errors: list[BulkUploadError] = []

    for excel_row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None or cell == "" for cell in row):
            continue

        raw = {h: row[i] if i < len(row) else None for h, i in idx.items()}

        try:
            payload = {
                "lb_pic_version": (
                    str(raw["lb_pic_version"]).strip()
                    if raw.get("lb_pic_version") not in (None, "")
                    else None
                ),
                "niveau": _normalize_niveau(raw.get("niveau")),
                "co_regate": _normalize_co_regate(raw.get("co_regate")),
                "dt_activation": _normalize_datetime(raw.get("dt_activation")),
                "dt_desactivation": _normalize_datetime(raw.get("dt_desactivation")),
                "motif_desactivation": (
                    str(raw["motif_desactivation"]).strip()
                    if raw.get("motif_desactivation") not in (None, "")
                    else None
                ),
                "commentaire": (
                    str(raw["commentaire"]).strip()
                    if raw.get("commentaire") not in (None, "")
                    else None
                ),
                "est_par_defaut": _normalize_bool(
                    raw.get("est_par_defaut"), default=False
                ),
            }
            pv = PicVersionCreate.model_validate(payload)
            valid.append(pv)
        except ValidationError as e:
            errors.append(BulkUploadError(row=excel_row_num, error=str(e), raw=raw))
        except (ValueError, TypeError) as e:
            errors.append(BulkUploadError(row=excel_row_num, error=str(e), raw=raw))

    return valid, errors


def pic_version_to_insert_params(p: PicVersionCreate) -> tuple:
    return (
        p.lb_pic_version,
        p.niveau.value,
        p.co_regate,
        p.dt_activation,
        p.dt_desactivation,
        p.motif_desactivation,
        p.commentaire,
        1 if p.est_par_defaut else 0,
    )


__all__ = [
    "EXPECTED_HEADERS",
    "REQUIRED_HEADERS",
    "INSERT_SQL",
    "NiveauEnum",
    "parse_excel_pic_versions",
    "pic_version_to_insert_params",
]

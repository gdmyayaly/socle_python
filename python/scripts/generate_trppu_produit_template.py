"""Génère le template Excel pour l'upload massif de trppu_produit.

Usage :
    python scripts/generate_trppu_produit_template.py
    python scripts/generate_trppu_produit_template.py --output path/to/file.xlsx
"""

from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "co_produit",
    "lb_produit",
    "dt_desactivation",
    "motif_desactivation",
]
EXAMPLE_ROWS = [
    ("OO", "Ordinaire Ouvert", None, None),
    ("LR", "Lettre Recommandée", None, None),
    ("PR", "Presse", date(2030, 12, 31), "Fin de contrat"),
]
COLUMN_WIDTHS = {
    "co_produit": 12,
    "lb_produit": 30,
    "dt_desactivation": 16,
    "motif_desactivation": 35,
}


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "produits"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[header]

    for row_idx, row in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if HEADERS[col_idx - 1] == "dt_desactivation" and value:
                cell.number_format = "YYYY-MM-DD"

    co_letter = get_column_letter(HEADERS.index("co_produit") + 1)
    for cell in ws[co_letter][1:]:
        cell.number_format = "@"

    ws.freeze_panes = "A2"

    notice = wb.create_sheet(title="notice")
    notice["A1"] = "Template d'import — table trppu_produit"
    notice["A1"].font = Font(bold=True, size=14)
    instructions = [
        "",
        "Colonnes :",
        " - co_produit          : code produit (2 caractères, obligatoire, clé primaire)",
        " - lb_produit          : libellé (max 80 caractères, obligatoire)",
        " - dt_desactivation    : date de désactivation (optionnelle, format YYYY-MM-DD)",
        " - motif_desactivation : motif (max 255 caractères, optionnel)",
        "",
        "Note : dt_creation est gérée automatiquement par la base (CURRENT_TIMESTAMP).",
        "",
        "Comportement de l'import :",
        " - Upsert sur co_produit (INSERT ... ON DUPLICATE KEY UPDATE).",
        " - Si le produit existe déjà, ses champs sont mis à jour.",
        " - Les lignes invalides sont retournées dans la réponse JSON sans bloquer le reste.",
        "",
        "Endpoint :",
        " - POST /trppu-api/produits/upload-excel  (multipart/form-data, champ 'file')",
    ]
    for i, line in enumerate(instructions, start=2):
        notice[f"A{i}"] = line
    notice.column_dimensions["A"].width = 110

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        "-o",
        default="trppu_produit_template.xlsx",
        help="Chemin du fichier Excel à générer (défaut : ./trppu_produit_template.xlsx)",
    )
    args = parser.parse_args()

    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(output_path)
    print(f"Template généré : {output_path}")


if __name__ == "__main__":
    main()

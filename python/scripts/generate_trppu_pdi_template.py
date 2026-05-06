"""Génère le template Excel pour l'upload massif de trppu_pdi.

Usage :
    python scripts/generate_trppu_pdi_template.py
    python scripts/generate_trppu_pdi_template.py --output path/to/file.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = ["id_pdi", "co_regate", "lb_pdi", "est_actif"]
EXAMPLE_ROWS = [
    (10001, "012345", "PDI Paris Centre", 1),
    (10002, "012345", "PDI Paris Sud", 1),
    (20001, "067890", "PDI Lyon Bellecour", 0),
]
COLUMN_WIDTHS = {"id_pdi": 14, "co_regate": 12, "lb_pdi": 40, "est_actif": 12}


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "pdis"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[header]

    for row_idx, row in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    co_letter = get_column_letter(HEADERS.index("co_regate") + 1)
    for cell in ws[co_letter][1:]:
        cell.number_format = "@"

    last_row = 1000
    actif_col = get_column_letter(HEADERS.index("est_actif") + 1)
    dv_actif = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error="Valeurs autorisées : 0 (inactif) ou 1 (actif)",
    )
    ws.add_data_validation(dv_actif)
    dv_actif.add(f"{actif_col}2:{actif_col}{last_row}")

    ws.freeze_panes = "A2"

    notice = wb.create_sheet(title="notice")
    notice["A1"] = "Template d'import — table trppu_pdi"
    notice["A1"].font = Font(bold=True, size=14)
    instructions = [
        "",
        "Colonnes :",
        " - id_pdi    : identifiant PDI (entier > 0, obligatoire, clé primaire)",
        " - co_regate : code Regate du site parent (6 caractères, obligatoire, doit exister dans trppu_site)",
        " - lb_pdi    : libellé (max 150 caractères, optionnel)",
        " - est_actif : 1 = actif (défaut), 0 = inactif",
        "",
        "Comportement de l'import :",
        " - Upsert sur id_pdi (INSERT ... ON DUPLICATE KEY UPDATE).",
        " - Pré-vérification que chaque co_regate existe dans trppu_site ; sinon ligne en erreur.",
        " - Les lignes invalides sont retournées dans la réponse JSON sans bloquer le reste.",
        "",
        "Endpoint :",
        " - POST /trppu-api/pdis/upload-excel  (multipart/form-data, champ 'file')",
    ]
    for i, line in enumerate(instructions, start=2):
        notice[f"A{i}"] = line
    notice.column_dimensions["A"].width = 110

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        "-o",
        default="trppu_pdi_template.xlsx",
        help="Chemin du fichier Excel à générer (défaut : ./trppu_pdi_template.xlsx)",
    )
    args = parser.parse_args()

    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(output_path)
    print(f"Template généré : {output_path}")


if __name__ == "__main__":
    main()

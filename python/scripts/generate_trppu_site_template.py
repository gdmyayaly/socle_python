"""Génère le template Excel pour l'upload massif de trppu_site.

Usage :
    python scripts/generate_trppu_site_template.py
    python scripts/generate_trppu_site_template.py --output path/to/file.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = ["co_regate", "lb_regate", "type_site", "co_roc"]
EXAMPLE_ROWS = [
    ("012345", "Regate Paris Nord", "PIC", "012345"),
    ("067890", "Regate Lyon Centre", "PDC1", "067890"),
    ("099999", "Regate Marseille",  "PPDC", "099999"),
]
COLUMN_WIDTHS = {
    "co_regate": 12,
    "lb_regate": 30,
    "type_site": 14,
    "co_roc": 12,
}


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "sites"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[header]

    for row_idx, row in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # co_regate / co_roc / type_site : format texte pour préserver les zéros initiaux
    # et éviter toute coercition Excel.
    for header in ("co_regate", "co_roc", "type_site"):
        col_letter = get_column_letter(HEADERS.index(header) + 1)
        for cell in ws[col_letter][1:]:
            cell.number_format = "@"

    ws.freeze_panes = "A2"

    notice = wb.create_sheet(title="notice")
    notice["A1"] = "Template d'import — table trppu_site"
    notice["A1"].font = Font(bold=True, size=14)
    instructions = [
        "",
        "Colonnes :",
        " - co_regate  : code Regate du site (6 caractères, obligatoire, clé primaire)",
        " - lb_regate  : libellé associé au code Regate (max 120 caractères, optionnel)",
        " - type_site  : type du site (max 5 caractères, ex : PIC, PDC1, PDC2, PPDC, AUTRE)",
        " - co_roc     : code ROC (6 caractères, obligatoire)",
        "",
        "Comportement de l'import :",
        " - Upsert sur co_regate (INSERT ... ON DUPLICATE KEY UPDATE).",
        " - Si le co_regate existe déjà, ses champs sont mis à jour.",
        " - Les lignes invalides sont retournées dans la réponse JSON sans bloquer le reste.",
        "",
        "Endpoint :",
        " - POST /trppu-api/sites/upload-excel  (multipart/form-data, champ 'file')",
    ]
    for i, line in enumerate(instructions, start=2):
        notice[f"A{i}"] = line
    notice.column_dimensions["A"].width = 110

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        "-o",
        default="trppu_site_template.xlsx",
        help="Chemin du fichier Excel à générer (défaut : ./trppu_site_template.xlsx)",
    )
    args = parser.parse_args()

    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(output_path)
    print(f"Template généré : {output_path}")


if __name__ == "__main__":
    main()

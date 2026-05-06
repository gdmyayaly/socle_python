"""Génère le template Excel pour l'upload massif de trppu_pic_coefficients.

Usage :
    python scripts/generate_trppu_pic_coefficients_template.py
    python scripts/generate_trppu_pic_coefficients_template.py --output path/to/file.xlsx
"""

from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "id_pic_version",
    "co_produit",
    "jour_semaine",
    "dt_effet",
    "dt_fin_effet",
    "coef_dense",
    "coef_faible1",
    "coef_faible2",
]
JOUR_VALUES = ["LUN", "MAR", "MER", "JEU", "VEN", "SAM"]
EXAMPLE_ROWS = [
    (1, "OO", "LUN", date(2026, 1, 1), None, 1.0500, 0.8000, 0.6000),
    (1, "OO", "MAR", date(2026, 1, 1), None, 1.0500, 0.8000, 0.6000),
    (1, "LR", "LUN", date(2026, 1, 1), date(2026, 12, 31), 1.2000, 0.9000, 0.7500),
]
COLUMN_WIDTHS = {
    "id_pic_version": 16,
    "co_produit": 12,
    "jour_semaine": 14,
    "dt_effet": 14,
    "dt_fin_effet": 14,
    "coef_dense": 12,
    "coef_faible1": 14,
    "coef_faible2": 14,
}


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "pic_coefficients"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[header]

    for row_idx, row in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            header = HEADERS[col_idx - 1]
            if header in ("dt_effet", "dt_fin_effet") and value:
                cell.number_format = "YYYY-MM-DD"
            elif header in ("coef_dense", "coef_faible1", "coef_faible2"):
                cell.number_format = "0.0000"

    co_letter = get_column_letter(HEADERS.index("co_produit") + 1)
    for cell in ws[co_letter][1:]:
        cell.number_format = "@"

    last_row = 1000

    jour_col = get_column_letter(HEADERS.index("jour_semaine") + 1)
    dv_jour = DataValidation(
        type="list",
        formula1='"' + ",".join(JOUR_VALUES) + '"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error=f"Valeurs autorisées : {', '.join(JOUR_VALUES)}",
    )
    ws.add_data_validation(dv_jour)
    dv_jour.add(f"{jour_col}2:{jour_col}{last_row}")

    ws.freeze_panes = "A2"

    notice = wb.create_sheet(title="notice")
    notice["A1"] = "Template d'import — table trppu_pic_coefficients"
    notice["A1"].font = Font(bold=True, size=14)
    instructions = [
        "",
        "Note : id_pic_coef n'est pas dans le template — auto-généré par la base.",
        "L'upload est un UPSERT sur la clé naturelle (id_pic_version, co_produit, jour_semaine, dt_effet) :",
        "  - si la combinaison n'existe pas → INSERT (nouvelle ligne)",
        "  - si elle existe déjà → UPDATE (dt_fin_effet et coefficients)",
        "",
        "Colonnes :",
        " - id_pic_version : id de la version PIC (entier > 0, obligatoire, doit exister dans trppu_pic_version)",
        " - co_produit     : code produit (2 caractères, obligatoire, doit exister dans trppu_produit)",
        " - jour_semaine   : LUN | MAR | MER | JEU | VEN | SAM (obligatoire)",
        " - dt_effet       : date de prise d'effet (obligatoire, format YYYY-MM-DD)",
        " - dt_fin_effet   : date de fin d'effet (optionnelle, doit être > dt_effet)",
        " - coef_dense     : coefficient zone dense (>= 0, max 4 décimales, obligatoire)",
        " - coef_faible1   : coefficient zone faible 1 (>= 0, max 4 décimales, obligatoire)",
        " - coef_faible2   : coefficient zone faible 2 (>= 0, max 4 décimales, obligatoire)",
        "",
        "Comportement de l'import :",
        " - Pré-vérification que chaque id_pic_version existe dans trppu_pic_version",
        " - Pré-vérification que chaque co_produit existe dans trppu_produit",
        " - Les lignes invalides sont retournées dans la réponse JSON sans bloquer le reste.",
        "",
        "Endpoint :",
        " - POST /trppu-api/pic-coefficients/upload-excel  (multipart/form-data, champ 'file')",
    ]
    for i, line in enumerate(instructions, start=2):
        notice[f"A{i}"] = line
    notice.column_dimensions["A"].width = 110

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        "-o",
        default="trppu_pic_coefficients_template.xlsx",
        help="Chemin du fichier Excel à générer (défaut : ./trppu_pic_coefficients_template.xlsx)",
    )
    args = parser.parse_args()

    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(output_path)
    print(f"Template généré : {output_path}")


if __name__ == "__main__":
    main()

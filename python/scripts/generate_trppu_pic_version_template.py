"""Génère le template Excel pour l'upload massif de trppu_pic_version.

Usage :
    python scripts/generate_trppu_pic_version_template.py
    python scripts/generate_trppu_pic_version_template.py --output path/to/file.xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "lb_pic_version",
    "niveau",
    "co_regate",
    "dt_activation",
    "dt_desactivation",
    "motif_desactivation",
    "commentaire",
    "est_par_defaut",
]
NIVEAU_VALUES = ["NATIONAL", "DEX", "SITE"]
EXAMPLE_ROWS = [
    ("PIC 2026 v1", "NATIONAL", "012345", datetime(2026, 1, 1, 0, 0, 0), None, None, "Première version", 1),
    ("PIC Paris Q2", "SITE", "012345", datetime(2026, 4, 1, 0, 0, 0), None, None, "Specifique IDF", 0),
    ("PIC DEX RP", "DEX", "067890", datetime(2026, 1, 1, 0, 0, 0), datetime(2026, 12, 31, 23, 59, 59), "Fin d'année", "DEX Rhône", 0),
]
COLUMN_WIDTHS = {
    "lb_pic_version": 22,
    "niveau": 12,
    "co_regate": 12,
    "dt_activation": 20,
    "dt_desactivation": 20,
    "motif_desactivation": 30,
    "commentaire": 30,
    "est_par_defaut": 14,
}


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "pic_versions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[header]

    for row_idx, row in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if HEADERS[col_idx - 1] in ("dt_activation", "dt_desactivation") and value:
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

    co_letter = get_column_letter(HEADERS.index("co_regate") + 1)
    for cell in ws[co_letter][1:]:
        cell.number_format = "@"

    last_row = 1000

    niveau_col = get_column_letter(HEADERS.index("niveau") + 1)
    dv_niveau = DataValidation(
        type="list",
        formula1='"' + ",".join(NIVEAU_VALUES) + '"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error=f"Valeurs autorisées : {', '.join(NIVEAU_VALUES)}",
    )
    ws.add_data_validation(dv_niveau)
    dv_niveau.add(f"{niveau_col}2:{niveau_col}{last_row}")

    defaut_col = get_column_letter(HEADERS.index("est_par_defaut") + 1)
    dv_defaut = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error="Valeurs autorisées : 0 ou 1",
    )
    ws.add_data_validation(dv_defaut)
    dv_defaut.add(f"{defaut_col}2:{defaut_col}{last_row}")

    ws.freeze_panes = "A2"

    notice = wb.create_sheet(title="notice")
    notice["A1"] = "Template d'import — table trppu_pic_version"
    notice["A1"].font = Font(bold=True, size=14)
    instructions = [
        "",
        "Note importante : id_pic_version n'est PAS dans le template.",
        "Cette colonne est auto-générée par la base à l'insertion.",
        "L'upload est INSERT-only : pour modifier une version existante, utiliser PUT /trppu-api/pic-versions/{id}.",
        "",
        "Colonnes :",
        " - lb_pic_version      : libellé court (max 80 caractères, optionnel)",
        " - niveau              : NATIONAL | DEX | SITE (obligatoire)",
        " - co_regate           : code Regate du site parent (6 caractères, obligatoire, doit exister dans trppu_site)",
        " - dt_activation       : date/heure d'activation (obligatoire, format YYYY-MM-DD HH:MM:SS)",
        " - dt_desactivation    : date/heure de désactivation (optionnelle, doit être > dt_activation)",
        " - motif_desactivation : motif (max 255 caractères, optionnel)",
        " - commentaire         : commentaire libre (max 500 caractères, optionnel)",
        " - est_par_defaut      : 1 = version par défaut pour le site, 0 = non (défaut)",
        "",
        "Comportement de l'import :",
        " - INSERT-only : chaque ligne crée une nouvelle version avec un id auto-généré.",
        " - Pré-vérification que chaque co_regate existe dans trppu_site ; sinon ligne en erreur.",
        " - Les lignes invalides sont retournées dans la réponse JSON sans bloquer le reste.",
        "",
        "Endpoint :",
        " - POST /trppu-api/pic-versions/upload-excel  (multipart/form-data, champ 'file')",
    ]
    for i, line in enumerate(instructions, start=2):
        notice[f"A{i}"] = line
    notice.column_dimensions["A"].width = 110

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        "-o",
        default="trppu_pic_version_template.xlsx",
        help="Chemin du fichier Excel à générer (défaut : ./trppu_pic_version_template.xlsx)",
    )
    args = parser.parse_args()

    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(output_path)
    print(f"Template généré : {output_path}")


if __name__ == "__main__":
    main()
